import openpyxl
from openpyxl.chart import BarChart, Reference, PieChart, Series
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_process_attack_board_template(filename="process_attack_board_template.xlsx"):
    """
    Creates an Excel template mimicking the process attack board image.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Process Attack Board"

    # Define styles
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Define KPIs and Sub-KPIs (placeholders)
    kpis = ["Downtime", "Part Fallout", "Change Points", "Quality Defects", "Process Struggles"]
    sub_kpis = {
        "Downtime": ["Sub-KPI 1", "Sub-KPI 2", "Sub-KPI 3"],
        "Part Fallout": ["Vendor NG", "Internal NG", "Scrap"],
        "Change Points": ["Engineering", "Process", "Material"],
        "Quality Defects": ["Type A", "Type B", "Type C"],
        "Process Struggles": ["Resource", "Method", "Measurement"],
    }

    # Header row
    header_row = ["KPI", "Sub-KPI", "PFS", "Value", "Notes"]
    for col_num, header in enumerate(header_row, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
        cell.fill = fill

    # Fill in KPI and Sub-KPI data
    row_num = 2
    for kpi in kpis:
        for sub_kpi in sub_kpis[kpi]:
            sheet.cell(row=row_num, column=1, value=kpi).border = border
            sheet.cell(row=row_num, column=2, value=sub_kpi).border = border
            sheet.cell(row=row_num, column=3, value="PFS").border = border
            sheet.cell(row=row_num, column=4, value="Value").border = border
            sheet.cell(row=row_num, column=5, value="Notes").border = border
            row_num += 1

    # Add charts (placeholders)
    # Example: Bar chart for Downtime
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Downtime Breakdown"
    chart1.x_axis.title = "Sub-KPI"
    chart1.y_axis.title = "Value"

    data = Reference(sheet, min_col=4, min_row=2, max_row=4) # Assuming Downtime Sub-KPIs are rows 2-4
    cats = Reference(sheet, min_col=2, min_row=2, max_row=4)

    chart1.add_data(data)
    chart1.set_categories(cats)
    sheet.add_chart(chart1, "G2") # Place chart in cell G2

    # Example: Pie chart for Part Fallout
    chart2 = PieChart()
    chart2.title = "Part Fallout Distribution"
    data2 = Reference(sheet, min_col=4, min_row=5, max_row=7) #Assuming Part Fallout Sub-KPIs are rows 5-7
    cats2 = Reference(sheet, min_col=2, min_row=5, max_row=7)
    series = Series(values=data2, title="Part Fallout")
    chart2.series = [series]
    chart2.set_categories(cats2)
    sheet.add_chart(chart2, "G18") # Place chart in cell G18

    # Adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    wb.save(filename)

if __name__ == "__main__":
    create_process_attack_board_template()
    print("Process Attack Board template created successfully!")